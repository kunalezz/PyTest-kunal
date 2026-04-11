import openpyxl

wb = openpyxl.Workbook()
sheetName = 'Sheet1'

if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'user'
ws['B1'] = 'password'

ws.append(['user1', '123'])
ws.append(['user2', '133'])
ws.append(['user3', '133'])
ws.append(['standard_user', 'secret_sauce'])
ws.append(['user4', '133'])

ws.delete_rows(3)

wb.save('sampleSheet.xlsx')

for i in ws.iter_rows(values_only=True):
    print(i)

